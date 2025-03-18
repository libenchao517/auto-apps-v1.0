################################################################################
# 本文件用于Excel表格的标准化
################################################################################
# 导入模块
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import Side
################################################################################
# 整理表格类
class Make_Table:
    def __init__(
            self,
            methods=None,
            is_bold_max = True,
            is_format_number = True
    ):
        """
        初始化函数
        :param methods: 用于排序的列表
        :param is_bold_max: 是否加粗最大值
        :param is_format_number: 是否格式化数字形式
        """
        self.methods = methods
        self.is_bold_max = is_bold_max
        self.is_format_number = is_format_number

    def Make(self, filename):
        """
        主函数
        :param filename: 文件路径和文件名
        :return: None
        """
        self.workbook = openpyxl.load_workbook(filename)
        self.sheets = self.workbook.sheetnames
        if self.methods is not None:
            self.sort_methods()
        self.set_font()
        if self.is_format_number:
            self.format_numbers()
        self.set_row_and_column()
        if self.is_bold_max:
            self.bold_max_values()
        self.set_center()
        self.set_three_line_table()
        self.workbook.save(filename)
        self.workbook.close()

    def format_numbers(self):
        """
        设置单元格格式为4位小数
        :return: None
        """
        for sheet_name in self.sheets:
            sheet = self.workbook[sheet_name]
            numeric_cells = [cell for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column) for cell in row if isinstance(cell.value, (float))]
            for cell in numeric_cells:
                cell.number_format = '#,##0.0000'

    def sort_methods(self):
        """
        如果列表self.methods非空，则按照self.methods的顺序对行进行排序
        对于不包含在self.methods中的行，放在表格最后
        :return: None
        """
        for sheet_name in self.sheets:
            sheet = self.workbook[sheet_name]
            data = list(sheet.iter_rows(min_row=2, values_only=True))
            title_row_index = False
            # 检查self.methods和表格第一列是否有重合
            for row_index, row in enumerate(data):
                if row[0] in self.methods:
                    title_row_index = True
                    break
            # 有重合时进行排序
            if title_row_index:
                sorted_data = sorted(data, key=lambda row : self.methods.index(row[0]) if row[0] in self.methods else float('inf'))
                sheet.delete_rows(sheet.min_row + 1, sheet.max_row)
                for row_data in sorted_data:
                    sheet.append(row_data)

    def set_row_and_column(self):
        """
        对第一行和第一列进行加粗
        :return: None
        """
        for sheet_name in self.sheets:
            sheet = self.workbook[sheet_name]
            # 加粗第一行
            for cell in sheet.iter_rows(min_row=1, max_row=1):
                for col_cell in cell:
                    col_cell.font = Font(bold=True, name='Times New Roman')
            # 加粗第一列
            for cell in sheet.iter_cols(min_col=1, max_col=1):
                for row_cell in cell:
                    row_cell.font = Font(bold=True, name='Times New Roman')

    def bold_max_values(self):
        """
        设置每一列的最大值为粗体
        :return: None
        """
        for sheet_name in self.sheets:
            sheet = self.workbook[sheet_name]
            for col in sheet.iter_cols(min_row=2, min_col=2, max_col=sheet.max_column):
                values = [cell.value for cell in col if cell.value is not None]  # 提取非空值
                if values:  # 如果该列有非空值，则加粗最大值
                    max_value = max(values)
                    for cell in col:
                        if cell.value == max_value:
                            cell.font = Font(bold=True, name='Times New Roman')

    def set_three_line_table(self):
        """
        设置表格为三线表
        :return: None
        """
        clear_border_style = Border()
        for sheet_name in self.sheets:
            sheet = self.workbook[sheet_name]
            # 先取消所有单元格的边框
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = clear_border_style
            # 设置第一行和最后一行的格式
            top_border_style = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'))
            bottom_border_style = Border(bottom=Side(border_style='medium'))
            for row_index in range(1, sheet.max_row + 1):
                for cell in sheet[row_index]:
                    if row_index == 1:
                        cell.border = top_border_style
                    elif row_index == sheet.max_row:
                        cell.border = bottom_border_style

    def set_center(self):
        """
        逐个工作簿、逐个单元格设置水平方向和垂直方向居中
        :return: None
        """
        for sheet_name in self.sheets:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(vertical='center', horizontal='center')

    def set_font(self):
        """
        逐个工作簿、逐个单元格设置字体为：Times New Roman
        :return: None
        """
        for sheet_name in self.sheets:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.font = Font(name='Times New Roman')
